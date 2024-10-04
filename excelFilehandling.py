import openpyxl
test=openpyxl.load_workbook('C:/Users/drpra/PycharmProjects/demoPytest/testdata/ReadingData.xlsx')
data=test['testData']
# print(data.max_row)
# print(data.max_column)
# print(data.cell(2,1))
print(data.cell(2,1).value)
print(data.cell(1,2).value)
